"""
BalanceLog Pro - Excel Exporter

Exports balancing records to formatted Excel files using openpyxl.
Creates professionally styled workbooks with headers, conditional formatting,
auto-column-width, and hyperlinks to screenshots.
"""

from pathlib import Path
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from src.database.models import BalancingRecord
from src.config.constants import EXCEL_HEADER_FILL_COLOR, EXCEL_HEADER_FONT_COLOR
from src.utils.logger import get_logger

logger = get_logger("excel")


# Column definitions: (header_text, field_name, width, number_format)
COLUMNS = [
    ("Date", "date", 14, None),
    ("Time", "time", 12, None),
    ("Punching Number", "punching_number", 20, None),
    ("Tube Length (mm)", "tube_length", 16, "0.0"),
    ("Type", "shaft_type", 10, None),
    ("Initial 0°", "initial_zero_degree", 12, "0.00"),
    ("Initial Left", "initial_left_value", 14, "0.00"),
    ("Initial Left Angle", "initial_left_angle", 16, "0.0"),
    ("Initial Right", "initial_right_value", 14, "0.00"),
    ("Initial Right Angle", "initial_right_angle", 16, "0.0"),
    ("Weight Add. Left", "weight_addition_left", 16, "0.00"),
    ("Weight Add. Right", "weight_addition_right", 16, "0.00"),
    ("After Corr. 0°", "after_correction_zero_degree", 14, "0.00"),
    ("After Corr. Left", "after_correction_left", 16, "0.00"),
    ("After Corr. Right", "after_correction_right", 16, "0.00"),
    ("Screenshot", "screenshot_path", 30, None),
    ("OCR Confidence", "ocr_confidence", 16, "0.0%"),
    ("Notes", "operator_notes", 25, None),
]


class ExcelExporter:
    """
    Exports balancing records to professionally formatted Excel workbooks.

    Features:
    - Styled header row with industrial dark blue theme
    - Auto-column-width
    - Number formatting for measurements
    - Conditional formatting for OCR confidence
    - Hyperlinks to screenshot files
    - Appends to existing daily files
    - Company title row
    """

    def __init__(self) -> None:
        # Header styles
        self._header_font = Font(
            name="Segoe UI", size=11, bold=True,
            color=EXCEL_HEADER_FONT_COLOR,
        )
        self._header_fill = PatternFill(
            start_color=EXCEL_HEADER_FILL_COLOR,
            end_color=EXCEL_HEADER_FILL_COLOR,
            fill_type="solid",
        )
        self._header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        self._header_border = Border(
            bottom=Side(style="thin", color="000000"),
            right=Side(style="thin", color="D0D0D0"),
        )

        # Data styles
        self._data_font = Font(name="Segoe UI", size=10)
        self._data_alignment = Alignment(horizontal="center", vertical="center")
        self._data_border = Border(
            bottom=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
        )

        # Title style
        self._title_font = Font(name="Segoe UI", size=14, bold=True, color="1E3A5F")

    # ─────────────────────────────────────────────────────────
    # Public Interface
    # ─────────────────────────────────────────────────────────
    def export_daily(self, date_str: str, records: List[BalancingRecord],
                     output_path: Path) -> Path:
        """
        Export records for a single day to an Excel file.

        If the file exists, appends new records. Otherwise creates a new file.

        Args:
            date_str: Date string (YYYY-MM-DD)
            records: List of balancing records to export
            output_path: Full path for the output Excel file

        Returns:
            Path to the created/updated Excel file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if output_path.exists():
            wb = load_workbook(str(output_path))
            ws = wb.active
            # Find last row and append
            start_row = ws.max_row + 1
            self._write_data_rows(ws, records, start_row)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Balancing Records {date_str}"
            self._write_title_row(ws, date_str)
            self._write_header_row(ws, row=3)
            self._write_data_rows(ws, records, start_row=4)
            self._set_column_widths(ws)

        wb.save(str(output_path))
        logger.info("Excel exported: %s (%d records)", output_path.name, len(records))
        return output_path

    def export_records(self, records: List[BalancingRecord],
                       output_path: Path, title: str = "Balancing Records") -> Path:
        """
        Export a list of records to a new Excel file (for search results, reports).

        Args:
            records: List of records to export
            output_path: Output file path
            title: Title for the worksheet

        Returns:
            Path to the created Excel file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name max 31 chars

        self._write_title_row(ws, title)
        self._write_header_row(ws, row=3)
        self._write_data_rows(ws, records, start_row=4)
        self._set_column_widths(ws)

        # Add summary row
        if records:
            self._write_summary_row(ws, records)

        wb.save(str(output_path))
        logger.info("Excel exported: %s (%d records)", output_path.name, len(records))
        return output_path

    # ─────────────────────────────────────────────────────────
    # Internal Writers
    # ─────────────────────────────────────────────────────────
    def _write_title_row(self, ws, title: str) -> None:
        """Write the company/title row at the top of the sheet."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        cell = ws.cell(row=1, column=1, value=f"BalanceLog Pro — {title}")
        cell.font = self._title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Subtitle with generation timestamp
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
        cell = ws.cell(
            row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )
        cell.font = Font(name="Segoe UI", size=9, italic=True, color="666666")
        cell.alignment = Alignment(horizontal="center")

    def _write_header_row(self, ws, row: int = 3) -> None:
        """Write the styled header row."""
        for col_idx, (header, _, _, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._header_alignment
            cell.border = self._header_border

        # Freeze panes below header
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

        # Auto-filter
        last_col = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A{row}:{last_col}{row}"

    def _write_data_rows(self, ws, records: List[BalancingRecord],
                         start_row: int) -> None:
        """Write data rows for all records."""
        for row_offset, record in enumerate(records):
            row = start_row + row_offset
            for col_idx, (_, field_name, _, num_format) in enumerate(COLUMNS, 1):
                value = getattr(record, field_name, "")
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = self._data_font
                cell.alignment = self._data_alignment
                cell.border = self._data_border

                if num_format:
                    cell.number_format = num_format

                # Screenshot column: add hyperlink
                if field_name == "screenshot_path" and value:
                    try:
                        cell.hyperlink = str(Path(value).as_uri())
                        cell.font = Font(
                            name="Segoe UI", size=10,
                            color="1E88E5", underline="single",
                        )
                        cell.value = Path(value).name  # Show filename only
                    except Exception:
                        pass

                # OCR confidence: conditional coloring
                if field_name == "ocr_confidence" and isinstance(value, (int, float)):
                    if value >= 0.90:
                        cell.font = Font(name="Segoe UI", size=10, color="43A047")
                    elif value >= 0.75:
                        cell.font = Font(name="Segoe UI", size=10, color="F9A825")
                    else:
                        cell.font = Font(name="Segoe UI", size=10, color="E53935", bold=True)

            # Alternate row shading
            if row_offset % 2 == 1:
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                    )

    def _set_column_widths(self, ws) -> None:
        """Set column widths based on definitions."""
        for col_idx, (_, _, width, _) in enumerate(COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    def _write_summary_row(self, ws, records: List[BalancingRecord]) -> None:
        """Write a summary row at the bottom of the data."""
        summary_row = ws.max_row + 2

        ws.cell(row=summary_row, column=1, value="Summary").font = Font(
            name="Segoe UI", size=11, bold=True,
        )
        ws.cell(row=summary_row, column=2, value=f"Total Records: {len(records)}").font = Font(
            name="Segoe UI", size=10,
        )

        if records:
            avg_confidence = sum(r.ocr_confidence for r in records) / len(records)
            ws.cell(
                row=summary_row, column=3,
                value=f"Avg OCR Confidence: {avg_confidence:.1%}",
            ).font = Font(name="Segoe UI", size=10)

            front_count = sum(1 for r in records if r.shaft_type == "Front")
            rear_count = sum(1 for r in records if r.shaft_type == "Rear")
            ws.cell(
                row=summary_row, column=4,
                value=f"Front: {front_count} | Rear: {rear_count}",
            ).font = Font(name="Segoe UI", size=10)
